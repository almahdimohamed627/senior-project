import os
from typing import List, Tuple

from openpyxl import load_workbook


def _safe_str(x) -> str:
    return ("" if x is None else str(x)).strip()


def load_raw_texts(data_dir: str = "data") -> List[Tuple[str, str]]:
    """
    Returns: List of (content, source_id)
    - .txt: (file_content, path)
    - .xlsx: each row becomes one "case" document:
        content: "التصنيف: ...\nوصف الحالة: ..."
        source_id: "path::sheet#row"
    """
    texts: List[Tuple[str, str]] = []

    if not os.path.isdir(data_dir):
        raise ValueError(f"مجلد البيانات '{data_dir}' مو موجود.")

    for root, _, files in os.walk(data_dir):
        for name in files:
            lower = name.lower()
            path = os.path.join(root, name)

            # ---- TXT ----
            if lower.endswith(".txt"):
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read().strip()
                    if content:
                        texts.append((content, path))
                continue

            # ---- XLSX ----
            if lower.endswith(".xlsx"):
                try:
                    wb = load_workbook(path, data_only=True, read_only=True)
                except Exception as e:
                    print(f"⚠️ Failed to read xlsx: {path} -> {e}")
                    continue

                for ws in wb.worksheets:
                    rows = ws.iter_rows(values_only=True)
                    header = next(rows, None)
                    if not header:
                        continue

                    header_norm = [_safe_str(h).lower() for h in header]

                    # Expect columns: Category, Description (case-insensitive)
                    try:
                        cat_i = header_norm.index("category")
                        desc_i = header_norm.index("description")
                    except ValueError:
                        # sheet ما فيه هالعناوين، تجاهله
                        continue

                    for r_idx, row in enumerate(rows, start=2):  # start=2 لأن header هو row 1
                        row = row or ()
                        cat = _safe_str(row[cat_i]) if cat_i < len(row) else ""
                        desc = _safe_str(row[desc_i]) if desc_i < len(row) else ""
                        if not desc:
                            continue

                        content = f"التصنيف: {cat}\nوصف الحالة: {desc}".strip()
                        source_id = f"{path}::{ws.title}#{r_idx}"
                        texts.append((content, source_id))

    if not texts:
        raise ValueError(f"ما لقيت ولا ملف (.txt أو .xlsx) فيه محتوى داخل '{data_dir}'.")

    return texts
